import os

import mysql.connector
import pandas as pd
import xlrd
from openpyxl import Workbook, load_workbook

from database_query import my_cursor, my_db, Database

global my_sheet


def createCompanySheet(hr, acc, pro, lead):
    try:
        # Code adapted from:
        add_employee = """INSERT INTO com_details (HR_Name, Acc_Manager_Name, Pro_Manager_Name, Lead_Name) 
                   VALUES (%s, %s, %s, %s) """

        dep_data = (hr, acc, pro, lead)
        my_cursor.execute(add_employee, dep_data)
        my_db.commit()

    except mysql.connector.Error as error:
        print("Failed to insert into MySQL table {}".format(error))


# createCompanySheet(1, "Tom", "Herbie", "Jazz", "Hermon")


def mergeCompanySheet():
    global my_sheet
    try:
        # Open existing excel file which you want to merge ------------------------
        other_workbook = input("Enter the file name in .xlsx format only: ")
        path = os.path.abspath(other_workbook)
        print(path)
        wb = load_workbook(str(other_workbook))
        print(wb.sheetnames)

        try:
            other_workbook_sheet = input("Enter the sheet name you want to merge\n")
            book = pd.read_excel(
                str(other_workbook), sheet_name=str(other_workbook_sheet)
            )
            book.drop_duplicates(
                subset=[
                    "Uni_ID",
                    "HR_Name",
                    "Acc_Manager_Name",
                    "Pro_Manager_Name",
                    "Lead_Name",
                ],
                inplace=True,
                keep="last",
            )

            wb = load_workbook("Company_Workbook.xlsx")
            print(wb.sheetnames)
            # wb.save("file.xls")
            my_sheet = input("In which sheet you want to merge\n")
            book.to_excel(
                "Company_Workbook.xlsx", sheet_name=str(my_sheet), index=False
            )

            # Open existing excel file which you want to merge ------------------------
            book = xlrd.open_workbook("Company_Workbook.xlsx")
            sheet = book.sheet_by_name(str(my_sheet))
            Database.dbQueries()
            Database.companyTableForm()

            try:

                # Code adapted from:
                add_employee = """INSERT INTO com_details (HR_Name, Acc_Manager_Name, Pro_Manager_Name, Lead_Name) 
                                           VALUES (%s, %s, %s, %s) """

                for r in range(1, sheet.nrows):
                    hr = sheet.cell(r, 1).value
                    acc = sheet.cell(r, 2).value
                    pro = sheet.cell(r, 3).value
                    lead = sheet.cell(r, 4).value

                    values = (hr, acc, pro, lead)
                    my_cursor.execute(add_employee, values)
                deleteDuplicateRows()
                my_db.commit()
                print("")
                print("All done! Bye, for now.")
                print("")
                columns = str(sheet.ncols)
                rows = str(sheet.nrows)
                print(f"I Just imported {columns} columns and {rows} rows to MySQL.")
                mergeCompanyWorkbook()

            except mysql.connector.Error as error:
                print("Failed to insert into MySQL table {}".format(error))

        except KeyError as error:
            print("Failed to open file {}".format(error))

    except FileNotFoundError as error:
        print("Failed to insert into MySQL table {}".format(error))


def saveCompanyWorkbook():
    global my_sheet
    # Create Excel (.xlsx) file -----------------------------------------------
    wb = Workbook()

    SQL = "SELECT * from " + "com_details" + ";"
    my_cursor.execute(SQL)
    results = my_cursor.fetchall()
    ws = wb.create_sheet(index=0)
    try:
        my_sheet = input(
            "Press 1 for default sheet name, Press 2 for you want to name the sheet\n"
        )
        if my_sheet == "1":
            ws.title = "com_details"
        elif my_sheet == "2":
            NewName = input("What Do you Want to Name the Sheet1 ?\n")
            ws.title = NewName
        ws.append(my_cursor.column_names)
        for row in results:
            ws.append(row)
        deleteDuplicateRows()
        workbook_name = "Company_Workbook"
        wb.save(workbook_name + ".xlsx")
    except mysql.connector.Error as error:
        print("Failed to insert into MySQL table {}".format(error))


def mergeCompanyWorkbook():
    global my_sheet
    try:
        # Create Excel (.xlsx) file -----------------------------------------------
        wb = Workbook()

        SQL = "SELECT * from " + "com_details" + ";"
        my_cursor.execute(SQL)
        results = my_cursor.fetchall()
        ws = wb.create_sheet(index=0)
        ws.title = my_sheet
        ws.append(my_cursor.column_names)
        for row in results:
            ws.append(row)
        deleteDuplicateRows()
        workbook_name = "Company_Workbook"
        wb.save(workbook_name + ".xlsx")

    except mysql.connector.Error as error:
        print("Failed to insert into MySQL table {}".format(error))


def deleteDuplicateRows():
    SQL = "SELECT * from " + "com_details" + ";"
    my_cursor.execute(SQL)
    my_cursor.fetchall()
    SQL = """DELETE t1 FROM com_details t1 INNER JOIN com_details t2 WHERE
    t1.Uni_ID < t2.Uni_ID and t1.HR_Name = t2.HR_Name and t1.Acc_Manager_Name = t2.Acc_Manager_Name and 
    t1.Pro_Manager_Name = t2.Pro_Manager_Name and t1.Lead_Name = t2.Lead_Name;"""
    my_cursor.execute(SQL)
    my_cursor.fetchall()
    SQL = """SET @num := 0;"""
    my_cursor.execute(SQL)
    SQL = """UPDATE com_details SET Uni_ID = @num := (@num+1);"""
    my_cursor.execute(SQL)
    SQL = """ALTER TABLE com_details AUTO_INCREMENT = 1;"""
    my_cursor.execute(SQL)
    my_cursor.fetchall()
    my_db.commit()


def deleteCompanyData():
    global my_sheet
    user = input(
        "Press 1 for find specific value and delete," "Press 2 for delete all data\n"
    )
    if user == "1":
        try:
            SQL = "SELECT * from " + "com_details" + ";"
            my_cursor.execute(SQL)
            my_cursor.fetchall()
            user = input(
                "Press 1 for delete by 'Uni_ID', "
                "Press 2 for delete by 'HR_Name', "
                "Press 3 for delete by 'Acc_Manager_Name', "
                "Press 4 for delete by 'Pro_Manager_Name', "
                "Press 5 for delete by 'Lead_Name'\n"
            )
            if user == "1":
                user = input("Enter the Uni_ID\n")
                SQL1 = f"""DELETE FROM com_details WHERE Uni_ID = '{user}'"""
                my_cursor.execute(SQL1)
                my_cursor.fetchall()
            elif user == "2":
                user = input("Enter the HR_Name\n")
                SQL2 = f"""DELETE FROM com_details WHERE HR_Name = '{user}'"""
                my_cursor.execute(SQL2)
                my_cursor.fetchall()
            elif user == "3":
                user = input("Enter the Acc_Manager_Name\n")
                SQL3 = f"""DELETE FROM com_details WHERE Acc_Manager_Name = '{user}'"""
                my_cursor.execute(SQL3)
                my_cursor.fetchall()
            elif user == "4":
                user = input("Enter the Pro_Manager_Name\n")
                SQL4 = f"""DELETE FROM com_details WHERE Pro_Manager_Name = '{user}'"""
                my_cursor.execute(SQL4)
                my_cursor.fetchall()
            elif user == "5":
                user = input("Enter the Lead_Name\n")
                SQL5 = f"""DELETE FROM com_details WHERE Lead_Name = '{user}'"""
                my_cursor.execute(SQL5)
                my_cursor.fetchall()
            my_db.commit()
            print("number of rows deleted", my_cursor.rowcount)
            deleteDuplicateRows()

        except mysql.connector.Error as error:
            print("Failed to delete data {}".format(error))

    elif user == "2":
        try:
            SQL = """TRUNCATE TABLE com_details"""
            my_cursor.execute(SQL)
            my_cursor.fetchall()
            my_db.commit()
            print("All Record Deleted successfully ")
            deleteDuplicateRows()

        except mysql.connector.Error as error:
            print("Failed to delete data {}".format(error))


def deleteInSheet():
    global my_sheet
    try:
        wb = load_workbook("Company_Workbook.xlsx")
        print(wb.sheetnames)
        my_sheet = input("In which sheet you want to use\n")

        # Open existing excel file which you want to merge ------------------------
        SQL = pd.read_sql("""SELECT * FROM com_details;""", my_db)
        SQL.to_excel("Company_Workbook.xlsx", sheet_name=str(my_sheet), index=False)

    except mysql.connector.Error as error:
        print("Failed to insert into MySQL table {}".format(error))
